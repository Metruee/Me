from __future__ import annotations
"""
Uploads API（文件上传与解析）
"""
import io
import json
import logging
import os
import re
import shutil
import subprocess
import threading
import time
import uuid as _uuid
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from core.config import UPLOAD_DIR, MAX_PARSE_CHARS

logger = logging.getLogger("me-backend")
router = APIRouter()

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# 文件解析类型映射
FILE_PARSERS = {
    ".pdf": "pdf", ".xlsx": "excel", ".xls": "excel", ".csv": "excel",
    ".png": "image", ".jpg": "image", ".jpeg": "image", ".webp": "image",
    ".txt": "text", ".md": "text", ".docx": "word",
}

# 解析任务缓存
_parsing_tasks = {}


def _clean_stale_tasks():
    now = time.time()
    stale = [tid for tid, t in _parsing_tasks.items()
             if now - t.get("started", 0) > 600]
    for tid in stale:
        _parsing_tasks.pop(tid, None)


def _parse_image_with_llm(file_path: Path) -> str:
    """用 LLM 多模态识别图片"""
    from core.config import LLM_BASE, LLM_MODEL
    db = None
    try:
        from core.db import get_db
        db = get_db()
        r1 = db.execute("SELECT value FROM settings WHERE key='llm_api_base'").fetchone()
        r2 = db.execute("SELECT value FROM settings WHERE key='llm_model'").fetchone()
        api_base = (r1["value"] if r1 else "") or LLM_BASE
        model = (r2["value"] if r2 else "") or LLM_MODEL
    except:
        api_base, model = LLM_BASE, LLM_MODEL
    finally:
        if db: db.close()

    import base64
    if model:
        try:
            b64 = base64.b64encode(file_path.read_bytes()).decode("utf-8")
            ext = file_path.suffix.lower().lstrip(".")
            if ext == "jpg":
                ext = "jpeg"
            data_uri = f"data:image/{ext};base64,{b64}"
            payload = {
                "model": model,
                "messages": [
                    {"role": "system", "content": "请详细描述图片中的内容。"},
                    {"role": "user", "content": [{"type": "image_url", "image_url": {"url": data_uri}}]},
                ],
                "max_tokens": 1024,
            }
            body = json.dumps(payload).encode("utf-8")
            req_url = f"{api_base.rstrip('/')}/chat/completions"
            req = __import__('urllib.request').request.Request(
                req_url, data=body,
                headers={"Content-Type": "application/json"},
                method="POST"
            )
            import urllib.request
            with urllib.request.urlopen(req, timeout=300) as resp:
                data = json.loads(resp.read())
                result = data["choices"][0]["message"]["content"]
                if result and result.strip():
                    return result.strip()
        except:
            pass

    # OCR 降级
    try:
        result = subprocess.run(
            ["tesseract", str(file_path), "stdout", "-l", "chi_sim+eng"],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0 and result.stdout.strip():
            return f"[OCR识别: {file_path.name}]\n{result.stdout.strip()[:3000]}"
    except:
        pass

    try:
        from PIL import Image
        img = Image.open(file_path)
        return f"[图片: {file_path.name}] 格式={img.format}, 尺寸={img.size[0]}×{img.size[1]}, 多模态不可用"
    except:
        pass
    return f"[图片: {file_path.name}] 无法识别"


def _parse_file(file_path: Path) -> str:
    ext = file_path.suffix.lower()
    parser_type = FILE_PARSERS.get(ext, "")

    if parser_type == "text" or ext in (".txt", ".md", ".csv"):
        return file_path.read_text(encoding="utf-8", errors="replace")[:MAX_PARSE_CHARS]

    if parser_type == "pdf":
        try:
            result = subprocess.run(
                ["pdftotext", "-layout", "-nopgbrk", str(file_path), "-"],
                capture_output=True, text=True, timeout=15
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout[:MAX_PARSE_CHARS]
        except:
            pass
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(file_path))
            text = "\n".join(p.extract_text() or "" for p in reader.pages)
            return text[:MAX_PARSE_CHARS] if text.strip() else ""
        except:
            pass
        return f"[PDF: {file_path.name}] 无法解析"

    if parser_type == "excel":
        try:
            import pandas as pd
            df = pd.read_excel(str(file_path)) if ext in (".xlsx", ".xls") else pd.read_csv(str(file_path))
            return df.to_string(max_rows=9999)[:MAX_PARSE_CHARS]
        except:
            pass
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(" | ".join(str(c or "") for c in row))
                parts.append(f"=== {sheet_name} ===\n" + "\n".join(rows[:9999]))
            return "\n\n".join(parts)[:MAX_PARSE_CHARS]
        except:
            pass
        return f"[Excel: {file_path.name}] 无法解析"

    if parser_type == "word":
        try:
            from docx import Document
            doc = Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs)[:MAX_PARSE_CHARS]
        except:
            pass
        return f"[Word: {file_path.name}] 无法解析"

    if parser_type == "image":
        return _parse_image_with_llm(file_path)

    try:
        return file_path.read_text(encoding="utf-8", errors="replace")[:3000]
    except:
        return f"[文件: {file_path.name}] 无法识别格式"


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    saved_name = file.filename or "upload"
    saved_path = UPLOAD_DIR / saved_name
    ext = saved_path.suffix.lower()
    if saved_path.exists():
        stem = saved_path.stem
        i = 1
        while (UPLOAD_DIR / f"{stem} ({i}){ext}").exists():
            i += 1
        saved_name = f"{stem} ({i}){ext}"
        saved_path = UPLOAD_DIR / saved_name
    with saved_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)

    task_id = f"parse_{_uuid.uuid4().hex[:12]}"
    _clean_stale_tasks()
    _parsing_tasks[task_id] = {"status": "parsing", "filename": saved_name, "started": time.time()}

    def _do_parse():
        try:
            text = _parse_file(saved_path)
            _parsing_tasks[task_id] = {
                "status": "done", "filename": saved_name,
                "parsed_text": text[:MAX_PARSE_CHARS],
                "parsed_length": len(text),
                "truncated": len(text) > MAX_PARSE_CHARS,
            }
        except Exception as e:
            logger.warning(f"Parse failed for {saved_name}: {e}")
            _parsing_tasks[task_id] = {"status": "error", "filename": saved_name, "error": str(e)}

    threading.Thread(target=_do_parse, daemon=True).start()
    return {"ok": True, "filename": saved_name, "task_id": task_id, "status": "parsing"}


@router.get("/upload/status/{task_id}")
def upload_status(task_id: str):
    task = _parsing_tasks.get(task_id)
    if not task:
        return {"ok": False, "error": "未知任务或已过期"}
    return {"ok": True, **{k: v for k, v in task.items() if k != "started"}}


@router.get("/uploads/{filename}")
def get_uploaded_file(filename: str):
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path)


@router.get("/uploads")
def list_uploads():
    if not UPLOAD_DIR.exists():
        return {"files": []}
    files = []
    for f in sorted(UPLOAD_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.is_file():
            st = f.stat()
            files.append({
                "name": f.name, "size": st.st_size,
                "mtime": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
                "ext": f.suffix.lower(),
            })
    return {"files": files}


@router.delete("/uploads/{filename}")
def delete_uploaded_file(filename: str):
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise HTTPException(404)
    path.unlink()
    return {"ok": True}
