"""
Me Backend — FastAPI

数据库：SQLite (7 张核心表)
向量库：ChromaDB（语义搜索 + 自动归档）

NAS 友好：单文件部署，< 50 MB 内存基线
"""
import json
import logging
import os
import re
import shutil
import sqlite3
import subprocess
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
import zipfile
import io
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import List, Optional, Union

from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("me-backend")

# ─── App ───────────────────────────────────────────────────
app = FastAPI(title="Me · 自知 API", version="0.3.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ─── Paths（支持环境变量覆盖，方便本地开发） ───────────────
ME_HOME     = Path(os.environ.get("ME_HOME", "/app/me_data"))
APP_ROOT    = Path(os.environ.get("ME_APP_ROOT", str(ME_HOME.parent)))
DB_PATH     = ME_HOME / "me.db"
REPORTS_DIR = APP_ROOT / "reports"
SKILLS_DIR  = Path(os.environ.get("SKILLS_DIR", str(ME_HOME / "skills")))
UPLOAD_DIR  = APP_ROOT / "data" / "uploads"
CHROMA_PATH = APP_ROOT / "chroma_data"
LLM_BASE    = os.environ.get("LLM_API_BASE", "http://192.168.1.100:11434/v1")
LLM_MODEL   = os.environ.get("LLM_MODEL", "")
DB_PATH.parent.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
CHROMA_PATH.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
MAX_PARSE_CHARS = 200000

# ─── ChromaDB + Embedding（延迟初始化）──────────────────────
_chroma_collection = None

def _get_chroma():
    global _chroma_collection
    if _chroma_collection is None:
        try:
            import chromadb
            client = chromadb.PersistentClient(
                path=str(CHROMA_PATH),
                settings=chromadb.config.Settings(anonymized_telemetry=False),
            )
            _chroma_collection = client.get_or_create_collection(
                name="me_knowledge", metadata={"hnsw:space": "cosine"},
            )
            logger.info(f"ChromaDB ready: {_chroma_collection.count()} entries")
        except Exception as e:
            logger.warning(f"ChromaDB unavailable: {e}")
            _chroma_collection = False
    return _chroma_collection if _chroma_collection is not False else None

def _embed(text: str) -> Optional[list]:
    """调用 Embedding API 生成向量"""
    db = get_db()
    try:
        row = db.execute("SELECT value FROM settings WHERE key='embedding_api_base'").fetchone()
        base = (row["value"] if row else "") or LLM_BASE
        row = db.execute("SELECT value FROM settings WHERE key='embedding_model'").fetchone()
        model = (row["value"] if row else "")
        if not base or not model: return None
        payload = json.dumps({"input": text, "model": model}).encode()
        req = urllib.request.Request(f"{base.rstrip('/')}/embeddings", data=payload,
            headers={"Content-Type":"application/json"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read())["data"][0]["embedding"]
    except Exception as e:
        logger.warning(f"Embedding failed: {e}")
        return None
    finally:
        db.close()

def _kb_search(query: str, domain: str = "", limit: int = 5) -> str:
    """知识库语义搜索 → 返回上下文文本，空字符串表示无结果"""
    coll = _get_chroma()
    entries = []
    # 1) ChromaDB 向量检索
    if coll:
        try:
            emb = _embed(query)
            if emb:
                where = None if (not domain or domain == "all") else {"theme_main": domain}
                r = coll.query(query_embeddings=[emb], n_results=limit, where=where,
                    include=["metadatas","documents","distances"])
                if r and r.get("documents") and r["documents"][0]:
                    for i, doc in enumerate(r["documents"][0]):
                        meta = r["metadatas"][0][i]
                        entries.append(f"[{meta.get('theme_main','')}] {doc[:200]}")
        except Exception as e:
            logger.warning(f"ChromaDB search failed: {e}")
    # 2) 退路：SQLite 关键词
    if not entries:
        db = get_db()
        try:
            words = [w for w in query.split() if len(w) > 1][:5]
            if words:
                patterns, params = [], []
                for w in words:
                    patterns.append("summary LIKE ?")
                    params.append(f"%{w}%")
                sql = f"SELECT theme_main, summary FROM knowledge_entries WHERE {' OR '.join(patterns)} ORDER BY created_at DESC LIMIT ?"
                rows = db.execute(sql, params + [limit]).fetchall()
                entries = [f"[{r['theme_main']}] {r['summary'][:200]}" for r in rows]
        finally:
            db.close()
    return "\n".join(f"- {e}" for e in entries) if entries else ""

def _kb_archive(entry_id: str, theme: str, summary: str, text: str):
    """知识条目 → ChromaDB 向量归档"""
    coll = _get_chroma()
    if not coll: return
    try:
        emb = _embed(summary or text[:500])
        if emb:
            coll.add(ids=[entry_id], embeddings=[emb], documents=[text],
                metadatas=[{"theme_main": theme or "未归类",
                    "summary": summary or text[:200],
                    "timestamp": datetime.now(timezone.utc).isoformat()}],
            )
    except Exception as e:
        logger.warning(f"ChromaDB archive failed: {e}")

# ─── SQLite ────────────────────────────────────────────────
def get_db() -> sqlite3.Connection:
    db = sqlite3.connect(str(DB_PATH))
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=ON")
    return db

def init_db():
    db = get_db()
    db.executescript("""
        CREATE TABLE IF NOT EXISTS experts (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            avatar TEXT DEFAULT '',
            system_prompt_file TEXT DEFAULT '',
            summon_phrase TEXT DEFAULT '',
            response_phrase TEXT DEFAULT '',
            domain TEXT DEFAULT 'all',
            is_enabled INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS conversations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            expert_id TEXT DEFAULT '',
            is_archived INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS knowledge_entries (
            id TEXT PRIMARY KEY,
            theme_main TEXT NOT NULL DEFAULT '',
            theme_sub TEXT DEFAULT '',
            summary TEXT DEFAULT '',
            original_text TEXT DEFAULT '',
            expert_id TEXT DEFAULT '',
            embedding_id TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS reports (
            id TEXT PRIMARY KEY,
            filename TEXT NOT NULL,
            period_start TEXT DEFAULT '',
            period_end TEXT DEFAULT '',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS skills (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL DEFAULT '',
            has_manifest INTEGER DEFAULT 0,
            enabled INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS sessions (
            id TEXT PRIMARY KEY,
            label TEXT NOT NULL DEFAULT '新对话',
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS daoben_entries (
            id TEXT PRIMARY KEY,
            event_text TEXT NOT NULL DEFAULT '',
            first_reaction TEXT DEFAULT '',
            greed TEXT DEFAULT '',
            fear TEXT DEFAULT '',
            excuses TEXT DEFAULT '',
            main_stone TEXT DEFAULT '',
            tomorrow_plan TEXT DEFAULT '',
            expert_id TEXT DEFAULT '',
            source TEXT DEFAULT 'manual',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS handoffs (
            id TEXT PRIMARY KEY,
            session_id TEXT NOT NULL,
            from_expert_id TEXT NOT NULL,
            to_expert_id TEXT NOT NULL,
            summary TEXT DEFAULT '',
            issue TEXT DEFAULT '',
            context TEXT DEFAULT '',
            used INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );
    """)
    # Seed default settings
    defaults = [
        ("llm_api_base", LLM_BASE),
        ("llm_model", LLM_MODEL),
        ("embedding_api_base", os.environ.get("EMBEDDING_API_BASE", LLM_BASE)),
        ("embedding_model", os.environ.get("EMBEDDING_MODEL", "text-embedding-bge-m3")),
        ("auto_archive", "true"),
        ("similarity_threshold", "0.6"),
        ("chat_history_rounds", "10"),
    ]
    for k, v in defaults:
        db.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", [k, v])
    # Seed experts (含 system_prompt_file 路径)
    experts_seed = [
        ("taishiling","太史令","📜","all","史官在侧，秉笔直书。","太史令在此。你的言行，我将如实录于竹帛。"),
        ("zhongkui","钟馗","⚔️","自我核心","三尺青锋，照我肝胆。","心中有鬼，方须照剑。你是来伏魔的，还是来求饶的？"),
        ("chiyou","蚩尤","🐉","事业","兵主旗下，雾散云开。","说敌情。畏刀避剑之人，不配站在我的旗下。"),
        ("bigan","比干","⚖️","财富","玲珑七窍，公断无私。","我无心，故不偏。把你那笔糊涂账，摊开来。"),
        ("tanlang","贪狼","🐺","人性","贪狼吞月，欲念昭然。","你身上每一寸欲望，都瞒不过我。说吧，这次想喂养哪一个？"),
        ("zaojun","司命灶君","🔥","亲密关系","灶火明堂，司命在场。","家宅之事，善恶功过，我记下了。从实道来。"),
        ("qibo","岐伯","🌿","健康","上古天真，问于天师。","身乃心之宅。你哪里失了调和，从实说来。"),
        ("cangjie","仓颉","🏺","自知","鸟兽蹄爪，皆有其迹。","你看到了什么？是河底的石头，还是水面上的波纹？"),
    ]
    for eid, name, avatar, domain, summon, resp in experts_seed:
        skill_file = f"{SKILLS_DIR}/me_experts/{eid}.md"
        db.execute(
            "INSERT OR IGNORE INTO experts(id,name,avatar,domain,summon_phrase,response_phrase,system_prompt_file) VALUES(?,?,?,?,?,?,?)",
            [eid, name, avatar, domain, summon, resp, skill_file]
        )
    db.commit()
    db.close()

init_db()


# ─── Models ────────────────────────────────────────────────

class ExpertOut(BaseModel):
    id: str; name: str; avatar: str; domain: str
    summon_phrase: str = ""; response_phrase: str = ""
    system_prompt: str = ""; is_enabled: bool = True

class ChatRequest(BaseModel):
    message: str
    expert_id: str = "taishiling"
    session_id: str = "default"
    is_file_upload: bool = False

class ChatResponse(BaseModel):
    reply: str
    expert_id: str
    switch_type: str = "none"
    system_message: str = ""
    archived: bool = False
    response_prefix: str = ""
    llm_error: str = ""

class DaobenEntryOut(BaseModel):
    id: str
    event_text: str = ""
    first_reaction: str = ""
    greed: str = ""
    fear: str = ""
    excuses: str = ""
    main_stone: str = ""
    tomorrow_plan: str = ""
    expert_id: str = ""
    source: str = "manual"
    created_at: str = ""


class KnowledgeEntryIn(BaseModel):
    theme_main: str = ""
    theme_sub: str = ""
    summary: str = ""
    original_text: str = ""
    expert_id: str = ""

class KnowledgeEntryOut(KnowledgeEntryIn):
    id: str; created_at: str = ""

class ReportOut(BaseModel):
    id: str; filename: str; created_at: str = ""
    period_start: str = ""; period_end: str = ""


# ─── Health ────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status":"ok","version":"0.2.0","timestamp":datetime.now(timezone.utc).isoformat()}


# ═══════════════════════════════════════════════════════════
#  Chat — 专家系统 + 知识库注入 + ChromaDB 向量检索
# ═══════════════════════════════════════════════════════════

# ─── 路由：全部由 LLM 意图分类决定 ────────────────────

# ─── 专家领域描述（LLM 意图分类用） ─────────────────
def _get_expert_domain_hints():
    """从数据库动态生成专家领域描述，新专家自动生效"""
    db = get_db()
    rows = db.execute(
        "SELECT id, name, domain FROM experts WHERE is_enabled=1 AND id != 'taishiling'"
    ).fetchall()
    db.close()
    lines = []
    for r in rows:
        if r["domain"]:
            lines.append(f"- {r['id']} ({r['name']}): {r['domain']}")
    lines.append("- taishiling (太史令): 无明显领域归属的日常对话")
    return "\n".join(lines)


def _get_valid_expert_ids():
    """从数据库动态获取所有启用的专家 ID"""
    db = get_db()
    rows = db.execute("SELECT id FROM experts WHERE is_enabled=1").fetchall()
    db.close()
    return [r["id"] for r in rows]


def _get_llm_config():
    """从 settings 表读取 LLM 配置，回退到环境变量"""
    db = get_db()
    row = db.execute("SELECT key, value FROM settings WHERE key IN ('llm_api_base','llm_model')").fetchall()
    db.close()
    cfg = {r["key"]: r["value"] for r in row}
    api_base = cfg.get("llm_api_base") or LLM_BASE
    model = cfg.get("llm_model") or LLM_MODEL
    return api_base.rstrip("/"), model


def _route_llm_intent(message: str) -> Optional[str]:
    """
    LLM 兜底意图分类：当关键词未命中时，用轻量 LLM 判断该找哪位专家。
    返回 expert_id 或 None（None 表示太史令自己回复）。
    """
    hints = _get_expert_domain_hints()
    valid_ids = _get_valid_expert_ids()
    prompt = f"""你是自知平台的智能分诊助手。用户说了一段话，你需要判断最适合由以下哪一位专家来回应。

{hints}

规则：
1. 如果用户的话有明显的情感倾诉、寻求分析、需要具体建议，选择最匹配的专家。
2. 如果用户只是寒暄、闲聊、无明确诉求，返回 taishiling。
3. 只返回 expert_id（一个单词），不要任何解释。

用户的话：{message[:300]}

expert_id:"""
    try:
        api_base, model = _get_llm_config()
        if not model:
            return None
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": "你是一个意图分类器。只输出一个单词的 expert_id，不要解释。"},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.1, "max_tokens": 20,
        }
        body = json.dumps(payload).encode("utf-8")
        llm_req = urllib.request.Request(
            f"{api_base}/chat/completions", data=body,
            headers={"Content-Type": "application/json"}, method="POST"
        )
        with urllib.request.urlopen(llm_req, timeout=300) as resp:
            data = json.loads(resp.read())
            result = data["choices"][0]["message"]["content"].strip().lower()
            # 从数据库动态获取有效 expert_id 列表
            valid_ids = _get_valid_expert_ids()
            for eid in valid_ids:
                if eid in result:
                    return eid
            return None
    except Exception:
        return None


# ─── 工具系统（联网搜索） ─────────────────────────────────
def _resolve_llm_config(db) -> tuple:
    """从 settings 表读取 LLM 配置"""
    api = LLM_BASE
    model = LLM_MODEL
    try:
        r1 = db.execute("SELECT value FROM settings WHERE key='llm_api_base'").fetchone()
        r2 = db.execute("SELECT value FROM settings WHERE key='llm_model'").fetchone()
        if r1 and r1["value"]:
            api = r1["value"]
        if r2 and r2["value"]:
            model = r2["value"]
    except:
        pass
    return api, model


def _get_enabled_tools() -> list:
    """获取已启用的技能 ID 列表"""
    tools = []
    try:
        if SKILLS_DIR.exists():
            for d in SKILLS_DIR.iterdir():
                if d.is_dir() and (d / "SKILL.md").exists():
                    sid = d.name
                    db2 = get_db()
                    row = db2.execute("SELECT enabled FROM skills WHERE id=?", [sid]).fetchone()
                    db2.close()
                    if not row or row["enabled"]:
                        tools.append(sid)
    except:
        pass
    return tools


def _get_tool_definitions() -> list:
    """构建工具定义：内置工具 + 已启用技能中声明 tools 的"""
    tool_defs = [
        {
            "type": "function",
            "function": {
                "name": "web_search",
                "description": "搜索互联网获取实时信息。当用户问及热点新闻、实时事件、最新资讯、事实查证时使用。返回搜索结果摘要列表。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "搜索关键词"},
                        "count": {"type": "integer", "description": "返回条数，默认 5, 最大 10", "default": 5},
                    },
                    "required": ["query"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "web_fetch",
                "description": "抓取指定 URL 的网页内容，提取正文文本。用于阅读搜索结果中的具体文章。",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string", "description": "要抓取的网页 URL"},
                        "max_chars": {"type": "integer", "description": "最大返回字符数，默认 4000", "default": 4000},
                    },
                    "required": ["url"],
                },
            },
        },
    ]
    # 扫描已启用技能，加载它们声明的工具
    for sid in _get_enabled_tools():
        fm = _read_skill_frontmatter(SKILLS_DIR / sid / "SKILL.md")
        if fm.get("tools"):
            for t in fm["tools"]:
                if isinstance(t, dict):
                    tool_defs.append({
                        "type": "function",
                        "function": t,
                    })
    return tool_defs


def _call_llm(base: str, model: str, messages: list, tools: Optional[list] = None) -> dict:
    """调用 LLM，返回 {"content": str, "tool_calls": [...]}
    超时/连接失败自动重试 1 次，400 错误精简历史后重试 1 次。
    """
    payload = {"model": model, "messages": messages, "temperature": 0.7, "max_tokens": 2048}
    if tools:
        payload["tools"] = tools
    body = json.dumps(payload).encode()
    req = urllib.request.Request(f"{base}/chat/completions", data=body,
        headers={"Content-Type": "application/json"})

    last_error = None
    for attempt in range(2):
        try:
            timeout = 180 if attempt == 0 else 300
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = json.loads(resp.read())
                msg = data["choices"][0]["message"]
                result = {"content": msg.get("content", ""), "tool_calls": msg.get("tool_calls")}
                logger.info(f"LLM call model={model} content_len={len(result.get('content','') or '')} tool_calls={len(result.get('tool_calls') or [])} finish_reason={data['choices'][0].get('finish_reason')}")
                return result
        except urllib.error.HTTPError as e:
            last_error = e
            if e.code == 400 and len(messages) > 5:
                logger.warning(f"LLM 400 error, retrying with minimal history")
                slim = [m for m in messages if m["role"] == "system"]
                last_user = None
                for m in reversed(messages):
                    if m["role"] == "user":
                        last_user = m
                        break
                if last_user:
                    slim.append(last_user)
                payload2 = {"model": model, "messages": slim, "temperature": 0.7, "max_tokens": 2048}
                body2 = json.dumps(payload2).encode()
                req2 = urllib.request.Request(f"{base}/chat/completions", data=body2,
                    headers={"Content-Type": "application/json"})
                try:
                    with urllib.request.urlopen(req2, timeout=300) as resp2:
                        data2 = json.loads(resp2.read())
                        msg2 = data2["choices"][0]["message"]
                        result = {"content": msg2.get("content", ""), "tool_calls": msg2.get("tool_calls")}
                        logger.info(f"LLM retry OK content_len={len(result.get('content','') or '')}")
                        return result
                except Exception:
                    pass
            break  # 非 400 或 400 但重试也失败，跳出
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            last_error = e
            if attempt == 0:
                logger.warning(f"LLM timeout/network error (attempt {attempt+1}/2), retrying...")
                continue
            break
        except Exception as e:
            last_error = e
            break

    raise last_error or RuntimeError("LLM call failed after retries")


def _execute_tool(name: str, args: dict) -> str:
    """执行工具调用：内置工具 + 技能注册的工具"""
    if name == "web_search":
        return _do_web_search(args.get("query", ""), args.get("count", 5))
    if name == "web_fetch":
        return _do_web_fetch(args.get("url", ""), args.get("max_chars", 4000))
    # 查找技能注册的 MCP 工具
    for sid in _get_enabled_tools():
        fm = _read_skill_frontmatter(SKILLS_DIR / sid / "SKILL.md")
        for t in (fm.get("tools") or []):
            if isinstance(t, dict) and t.get("name") == name:
                handler = t.get("handler")
                if handler == "web_search":
                    return _do_web_search(args.get("query", ""), args.get("count", 5))
                if handler == "web_fetch":
                    return _do_web_fetch(args.get("url", ""), args.get("max_chars", 4000))
    return f"未知工具: {name}"


def _do_web_search(query: str, count: int = 5) -> str:
    """联网搜索，返回格式化结果摘要"""
    engine = "duckduckgo"
    engines = {
        "google": "https://www.google.com/search?q=",
        "bing": "https://www.bing.com/search?q=",
        "baidu": "https://www.baidu.com/s?wd=",
        "duckduckgo": "https://html.duckduckgo.com/html/?q=",
    }
    url = engines.get(engine, engines["duckduckgo"]) + urllib.parse.quote(query)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
        html = urllib.request.urlopen(req, timeout=10).read().decode("utf-8", errors="ignore")
        text = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<style[^>]*>.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        if any(kw in text[:200].lower() for kw in ["captcha", "robot", "verify", "blocked", "javascript", "cookie", "enable js"]):
            return f"[搜索被拦截，建议换用百度搜索引擎重试]"
        # 只取前部分有意义的内容
        return f"【搜索结果】关键词：{query}\n\n{text[:3000]}"
    except Exception as e:
        return f"搜索失败: {e}"


def _do_web_fetch(url: str, max_chars: int = 4000) -> str:
    """抓取网页内容，提取正文"""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
        html = urllib.request.urlopen(req, timeout=10).read().decode("utf-8", errors="ignore")
        text = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<style[^>]*>.*?</style>", " ", text, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        return f"【网页内容】{url}\n\n{text[:max_chars]}"
    except Exception as e:
        return f"抓取失败: {e}"


def _chat_with_tools(base: str, model: str, messages: list) -> str:
    """带工具调用循环的对话"""
    # 第一轮：带工具定义
    result = _call_llm(base, model, messages, _get_tool_definitions())
    for i in range(3):  # 最多 3 轮
        logger.info(f"tool round={i} content_len={len(result.get('content','') or '')} tool_calls={len(result.get('tool_calls') or [])}")
        if result.get("content"):
            return result["content"]
        tool_calls = result.get("tool_calls")
        if not tool_calls:
            break
        for tc in tool_calls:
            fn = tc["function"]
            try:
                tool_result = _execute_tool(fn["name"], json.loads(fn["arguments"]))
            except Exception as e:
                tool_result = f"工具执行失败: {e}"
            logger.info(f"tool name={fn.get('name')} args={fn.get('arguments')} result_len={len(tool_result or '')}")
            messages.append({
                "role": "assistant",
                "tool_calls": [tc],
            })
            messages.append({
                "role": "tool",
                "tool_call_id": tc["id"],
                "content": tool_result,
            })
        # 后续轮：不带工具（避免死循环）
        result = _call_llm(base, model, messages, None)
    final = result.get("content", "") or "抱歉，搜索后仍无法回答。"
    logger.info(f"tool final content_len={len(final)}")
    return final



# ─── Expert Handoff ────────────────────────────────────────
def _generate_handoff(db, session_id, from_expert, to_expert, recent_msgs, user_message):
    """专家切换时 LLM 生成交接备忘录"""
    llm_api_base, llm_model = _resolve_llm_config(db)
    if not llm_model:
        return None
    history = recent_msgs[-6:] if len(recent_msgs) > 6 else recent_msgs
    history_text = "\n".join(
        f"{'用户' if m['role']=='user' else from_expert.get('name','')}: {m['content'][:300]}"
        for m in history
    )
    prompt = (
        f"你是「{from_expert.get('name','')}」（{from_expert.get('title','')}），正在把对话交接给「{to_expert.get('name','')}」（{to_expert.get('title','')}」。\n\n"
        "请用JSON输出，包含三个字段：summary（一句话概括≤50字）、issue（需解决的问题≤200字）、context（背景信息≤300字）。\n\n"
        f"最近对话：\n{history_text}\n\n用户最新：{user_message[:200]}\n仅输出JSON。"
    )
    try:
        result = _call_llm(llm_api_base, llm_model, [{"role": "user", "content": prompt}])
        content = result.get("content", "")
        m = re.search(r'\{[\s\S]*\}', content)
        if m:
            data = json.loads(m.group())
            return {
                "summary": (data.get("summary") or "")[:100],
                "issue": (data.get("issue") or "")[:500],
                "context": (data.get("context") or "")[:1000],
            }
    except Exception as e:
        logger.warning(f"Handoff LLM failed: {e}")
    return {
        "summary": f"用户从「{from_expert.get('name','')}」转来",
        "issue": user_message[:200],
        "context": f"此前由 {from_expert.get('name','')} 负责此话题。",
    }

def _get_unread_handoff(db, session_id, expert_id):
    row = db.execute(
        "SELECT * FROM handoffs WHERE session_id=? AND to_expert_id=? AND used=0 ORDER BY created_at DESC LIMIT 1",
        [session_id, expert_id]
    ).fetchone()
    return dict(row) if row else None

class SessionHistory(BaseModel):
    session_id: str
    messages: list

@app.get("/api/history")
def get_history(session_id: str = "default", limit: int = 100):
    """获取会话历史消息"""
    db = get_db()
    rows = db.execute(
        "SELECT role, content, expert_id FROM conversations WHERE session_id=? ORDER BY id ASC LIMIT ?",
        [session_id, limit]
    ).fetchall()
    db.close()
    return [dict(r) for r in rows]

@app.get("/api/sessions")
def list_sessions():
    """获取所有会话列表"""
    db = get_db()
    # 确保 default 会话存在（兼容旧数据）
    has_default = db.execute("SELECT id FROM sessions WHERE id='default'").fetchone()
    if not has_default:
        has_conv = db.execute("SELECT id FROM conversations WHERE session_id='default' LIMIT 1").fetchone()
        if has_conv:
            db.execute("INSERT INTO sessions(id, label) VALUES('default','默认对话')")
            db.commit()
    rows = db.execute("SELECT id, label, created_at, updated_at FROM sessions ORDER BY updated_at DESC").fetchall()
    db.close()
    return [{
        "id": r["id"],
        "label": r["label"],
        "date": r["updated_at"][:10] if r["updated_at"] else r["created_at"][:10],
    } for r in rows]

@app.post("/api/sessions")
def create_session():
    """创建新会话"""
    import uuid
    sid = f"session_{uuid.uuid4().hex[:12]}"
    db = get_db()
    db.execute("INSERT INTO sessions(id, label) VALUES(?,?)", [sid, "新对话"])
    db.commit()
    db.close()
    return {"id": sid, "label": "新对话"}

@app.put("/api/sessions/{session_id}")
def rename_session(session_id: str, body: dict):
    """重命名会话"""
    db = get_db()
    db.execute("UPDATE sessions SET label=?, updated_at=datetime('now') WHERE id=?",
               [body.get("label", "新对话"), session_id])
    db.commit()
    db.close()
    return {"ok": True}

@app.delete("/api/session/{session_id}")
def delete_session(session_id: str):
    """删除整个会话及其所有消息"""
    db = get_db()
    db.execute("DELETE FROM conversations WHERE session_id=?", [session_id])
    db.execute("DELETE FROM handoffs WHERE session_id=?", [session_id])
    db.execute("DELETE FROM sessions WHERE id=?", [session_id])
    db.commit()
    db.close()
    return {"ok": True}

@app.post("/api/chat")
def chat(req: ChatRequest):
    """
    对话接口：召唤检测 → 关键词路由 → LLM（带历史上下文）。
    支持流式 / 非流式自动降级。
    """
    db = get_db()
    # 确保会话记录存在
    db.execute("INSERT OR IGNORE INTO sessions(id, label) VALUES(?,?)", [req.session_id, "新对话"])
    db.execute("UPDATE sessions SET updated_at=datetime('now') WHERE id=?", [req.session_id])

    # 读取当前 expert 的人格文件
    expert = db.execute("SELECT * FROM experts WHERE id=?", [req.expert_id]).fetchone()
    expert_name = expert["name"] if expert else "太史令"
    skill_file = expert["system_prompt_file"] if expert and expert["system_prompt_file"] else ""
    system_prompt = "你是自知平台的专家。"
    if skill_file:
        try:
            system_prompt = Path(skill_file).read_text(encoding="utf-8")
        except Exception:
            pass

    # ── 1. 召唤语检测 ──────────────────────────────
    experts_rows = db.execute("SELECT * FROM experts WHERE is_enabled=1").fetchall()
    for row in experts_rows:
        summon = row["summon_phrase"] or ""
        if summon and summon in req.message and row["id"] != req.expert_id:
            # 生成交接备忘录
            handoff_id = f"ho_{uuid.uuid4().hex[:12]}"
            from_exp = db.execute("SELECT * FROM experts WHERE id=?", [req.expert_id]).fetchone()
            to_exp_row = db.execute("SELECT * FROM experts WHERE id=?", [row["id"]]).fetchone()
            prev = db.execute(
                "SELECT role, content FROM conversations WHERE session_id=? AND expert_id=? ORDER BY id DESC LIMIT 6",
                [req.session_id, req.expert_id]
            ).fetchall()
            prev = list(reversed(prev))
            hd = _generate_handoff(
                db, req.session_id,
                dict(from_exp) if from_exp else {"name": req.expert_id, "title": ""},
                dict(to_exp_row) if to_exp_row else {"name": row["id"], "title": ""},
                [dict(m) for m in prev], req.message,
            )
            if hd:
                db.execute(
                    "INSERT INTO handoffs(id,session_id,from_expert_id,to_expert_id,summary,issue,context) VALUES(?,?,?,?,?,?,?)",
                    [handoff_id, req.session_id, req.expert_id, row["id"],
                     hd.get("summary",""), hd.get("issue",""), hd.get("context","")]
                )
            db.execute("INSERT INTO conversations(session_id,role,content,expert_id) VALUES(?,?,?,?)",
                       [req.session_id, "user", req.message, req.expert_id])
            db.commit(); db.close()
            return ChatResponse(
                reply=row["response_phrase"] or "",
                expert_id=row["id"], switch_type="summon",
                system_message=f"✨ 召唤 {row['name']}...",
                response_prefix=row["response_phrase"] or "",
            )

    # ── 2. LLM 意图分类路由 ──────────────────────────
    # 对太史令的消息（>10字），用 LLM 判断该交给哪位专家
    auto_expert_id = None
    auto_system_message = ""
    if req.expert_id == "taishiling" and len(req.message) > 10:
        auto_expert_id = _route_llm_intent(req.message)
    if auto_expert_id and req.expert_id == "taishiling":
        target = db.execute("SELECT * FROM experts WHERE id=?", [auto_expert_id]).fetchone()
        if target:
            # 生成交接备忘录
            handoff_id = f"ho_{uuid.uuid4().hex[:12]}"
            from_exp = db.execute("SELECT * FROM experts WHERE id=?", [req.expert_id]).fetchone()
            prev = db.execute(
                "SELECT role, content FROM conversations WHERE session_id=? AND expert_id=? ORDER BY id DESC LIMIT 6",
                [req.session_id, req.expert_id]
            ).fetchall()
            prev = list(reversed(prev))
            hd = _generate_handoff(
                db, req.session_id,
                dict(from_exp) if from_exp else {"name": req.expert_id, "title": ""},
                dict(target) if target else {"name": auto_expert_id, "title": ""},
                [dict(m) for m in prev], req.message,
            )
            if hd:
                db.execute(
                    "INSERT INTO handoffs(id,session_id,from_expert_id,to_expert_id,summary,issue,context) VALUES(?,?,?,?,?,?,?)",
                    [handoff_id, req.session_id, req.expert_id, auto_expert_id,
                     hd.get("summary",""), hd.get("issue",""), hd.get("context","")]
                )
            db.execute("INSERT INTO conversations(session_id,role,content,expert_id) VALUES(?,?,?,?)",
                       [req.session_id, "user", req.message, req.expert_id])
            auto_system_message = f"太史令判断，转由 {target['name']} 回应"
            # 切换到目标专家，继续走下面的 LLM 调用流程（不提前 return）
            req.expert_id = auto_expert_id
            expert = target
            expert_name = expert["name"]
            skill_file = expert["system_prompt_file"] if expert["system_prompt_file"] else ""
            system_prompt = "你是自知平台的专家。"
            if skill_file:
                try:
                    system_prompt = Path(skill_file).read_text(encoding="utf-8")
                except Exception:
                    pass
    # ── 3. 构建对话历史（按专家隔离）────────────────────
    history_rounds = int(os.environ.get("CHAT_HISTORY_ROUNDS", "5"))
    try:
        val = db.execute("SELECT value FROM settings WHERE key='chat_history_rounds'").fetchone()
        if val:
            history_rounds = int(val["value"])
    except:
        pass

    # 当前专家的独立历史
    prev_msgs = db.execute(
        "SELECT role, content FROM conversations WHERE session_id=? AND expert_id=? ORDER BY id DESC LIMIT ?",
        [req.session_id, req.expert_id, history_rounds * 2]
    ).fetchall()
    prev_msgs = list(reversed(prev_msgs))

    # 如果刚切换专家，附带上一位专家的最后一条消息作为交接上下文
    handoff_context = ""
    handoff_row = _get_unread_handoff(db, req.session_id, req.expert_id)
    if handoff_row:
        handoff_context = f"【交接备忘录】\n上一专家：{handoff_row['from_expert_id']}\n核心问题：{handoff_row['issue']}\n背景：{handoff_row['context']}\n请在此基础上继续为用户服务。"
        db.execute("UPDATE handoffs SET used=1 WHERE id=?", [handoff_row["id"]])
        db.commit()

    messages = [{"role": "system", "content": system_prompt + ("\n\n" + handoff_context if handoff_context else "")}]
    for pm in prev_msgs:
        messages.append({"role": pm["role"], "content": pm["content"]})

    # ── 3.5 知识库语义搜索 + 注入上下文 ──────────────────
    kb_context = _kb_search(req.message, domain=expert["domain"] or "all", limit=5)
    # 专家名单：从 DB 动态生成，注入为知识上下文
    expert_rows = db.execute("SELECT id, name, domain FROM experts WHERE is_enabled=1").fetchall()
    expert_list_str = "\n".join(f"- {r['id']}（{r['name']}）: {r['domain']}" for r in expert_rows)
    knowledge_blocks = []
    if kb_context:
        knowledge_blocks.append(f"## 知识库记忆（与此话题相关的历史记录）\n{kb_context}")
    knowledge_blocks.append(f"## 系统专家名单（仅此 {len(expert_rows)} 位，不存在其他专家）\n{expert_list_str}")
    # 仓颉额外注入道痕上下文
    if req.expert_id == "cangjie":
        daoben_rows = db.execute(
            "SELECT main_stone, event_text, greed, fear, excuses, tomorrow_plan, created_at FROM daoben_entries ORDER BY created_at DESC LIMIT 30"
        ).fetchall()
        if daoben_rows:
            daoben_text = "## 用户道痕记录（最近30条，按时间倒序）\n" + "\n".join(
                f"- [{r['created_at'][:10]}] 🪨{r['main_stone'] or '未命名'} | 事件:{r['event_text'][:60]} | 贪:{r['greed'][:40] or '—'} | 惧:{r['fear'][:40] or '—'} | 借口:{r['excuses'][:40] or '—'} | 下次:{r['tomorrow_plan'][:40] or '—'}"
                for r in daoben_rows
            )
            # 统计重复石头
            stone_freq = {}
            for r in daoben_rows:
                s = (r["main_stone"] or "").strip()
                if s:
                    stone_freq[s] = stone_freq.get(s, 0) + 1
            if stone_freq:
                top = sorted(stone_freq.items(), key=lambda x: x[1], reverse=True)[:5]
                daoben_text += "\n\n### 高频心石 Top 5\n" + "\n".join(f"- {k} x{v}" for k, v in top)
                daoben_text += f"\n\n（总计 {len(daoben_rows)} 条道痕，{len(stone_freq)} 种心石）"
            knowledge_blocks.append(daoben_text)
    if knowledge_blocks:
        messages.insert(0, {
            "role": "system",
            "content": "\n\n".join(knowledge_blocks),
        })

    messages.append({"role": "user", "content": req.message})

    # ══ 4. 调用 LLM ═══════════════════════════════
    reply = ""
    llm_error = ""
    try:
        llm_api_base, llm_model = _resolve_llm_config(db)
        if not llm_model:
            raise Exception("LLM 模型未配置")
        llm_base = llm_api_base.rstrip("/")
        llm_reply = _chat_with_tools(llm_base, llm_model, messages)
        if llm_reply:
            reply = llm_reply
        else:
            logger.warning(f"LLM returned empty reply for expert={req.expert_id} session={req.session_id}")
            llm_error = "LLM 返回为空"
    except Exception as e:
        logger.warning(f"LLM chat failed: {e}")
        llm_error = "模型响应超时或不可用，已自动重试，请稍后再试"

    # ── 5. 存对话记录 ──────────────────────────────
    db.execute("INSERT INTO conversations(session_id,role,content,expert_id) VALUES(?,?,?,?)",
               [req.session_id, "user", req.message, req.expert_id])
    if reply and not llm_error:
        db.execute("INSERT INTO conversations(session_id,role,content,expert_id) VALUES(?,?,?,?)",
                   [req.session_id, "assistant", reply, req.expert_id])

    # ── 6. 智能归档判断 ────────────────────────────
    archived = False
    archive_system_msg = ""
    auto_archive = True
    try:
        aa = db.execute("SELECT value FROM settings WHERE key='auto_archive'").fetchone()
        if aa:
            auto_archive = aa["value"].lower() in ("true", "1", "yes", "on")
    except:
        pass

    should_archive = False
    archive_reason = ""

    # 触发条件 1：用户主动表达归档意图
    if auto_archive and _detect_archive_intent(req.message):
        should_archive = True
        archive_reason = "user_intent"

    # 触发条件 2：当前专家对话轮次 ≥ 20 轮
    if auto_archive and not should_archive:
        rounds = _count_expert_rounds(db, req.session_id, req.expert_id)
        if rounds >= 20:
            should_archive = True
            archive_reason = "rounds_threshold"
            archive_system_msg = f"📌 当前话题已讨论 {rounds} 轮，为你提炼归档..."

    # 触发条件 3：文件导入消息（强制归档）
    if req.is_file_upload:
        should_archive = True
        archive_reason = "file_upload"

    if should_archive:
        # 用 LLM 提炼摘要+主题
        archive_result = _llm_archive_summarize(db, req.session_id, req.expert_id)
        if archive_result:
            theme = archive_result.get("theme", "未归类")
            summary = archive_result.get("summary", req.message[:200])
        else:
            theme = _classify_theme(req.message)
            summary = req.message[:200]
        # 去重检查
        if not _dedup_check(db, summary):
            entry_id = f"me_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S%f')}"
            full_text = archive_result.get("full_text", req.message) if archive_result else req.message
            db.execute(
                "INSERT INTO knowledge_entries(id,theme_main,summary,original_text,expert_id) VALUES(?,?,?,?,?)",
                [entry_id, theme, summary, full_text, req.expert_id]
            )
            db.commit()
            t = threading.Thread(target=_kb_archive, args=(entry_id, theme, summary, full_text))
            t.start()
            archived = True
            if archive_reason == "user_intent":
                archive_system_msg = f"✅ 已归档到「{theme}」"
            elif archive_reason == "file_upload":
                archive_system_msg = f"📥 文件内容已归档到「{theme}」"

    db.commit(); db.close()
    switch_type = "auto" if auto_expert_id else "none"
    return ChatResponse(
        reply=reply, expert_id=req.expert_id, switch_type=switch_type,
        system_message=archive_system_msg or auto_system_message,
        archived=archived, llm_error=llm_error,
    )


def _fallback_reply(expert_id: str, message: str) -> str:
    """LLM 不可用时的本地关键词回复"""
    replies = {
        "taishiling": "录毕。是否需召其他专家共议？",
        "zhongkui": "这句话背后，藏着什么你不敢直面的真相？",
        "chiyou": "战场瞬息万变。说明你的下一步。",
        "bigan": "数字不会说谎。让我们看看这笔账。",
        "tanlang": "你渴求的，究竟是目标本身，还是它带来的幻象？",
        "zaojun": "家宅安宁，在于秩序。如实说来。",
        "qibo": "身乃心之宅。你哪里失了调和？",
    }
    return replies.get(expert_id, "请继续说。")


def _detect_archive_intent(message: str) -> bool:
    """检测用户是否显式表达了归档意图：归档/收藏/记录/保存/记下来"""
    keywords = ["归档", "收藏", "记录一下", "记下来", "保存", "存档", "存一下", "收录"]
    for kw in keywords:
        if kw in message:
            return True
    return False


def _count_expert_rounds(db, session_id: str, expert_id: str) -> int:
    """统计当前会话中当前专家的连续对话轮次（用户消息次数）"""
    rows = db.execute(
        "SELECT role FROM conversations WHERE session_id=? AND expert_id=? ORDER BY id DESC LIMIT 40",
        [session_id, expert_id]
    ).fetchall()
    user_count = 0
    for r in rows:
        if r["role"] == "user":
            user_count += 1
        else:
            break
    return user_count


def _dedup_check(db, text: str) -> bool:
    """检查最近知识条目中是否有高度相似的内容。返回True表示重复。"""
    recent = db.execute(
        "SELECT summary, original_text FROM knowledge_entries ORDER BY created_at DESC LIMIT 20"
    ).fetchall()
    if not recent:
        return False
    words = set(text[:200].split())
    for r in recent:
        existing = (r["summary"] or "") + (r["original_text"] or "")[:500]
        existing_words = set(existing.split())
        if existing_words and words:
            overlap = len(words & existing_words) / max(len(words | existing_words), 1)
            if overlap > 0.6:
                return True
    return False


def _llm_archive_summarize(db, session_id: str, expert_id: str) -> dict:
    """调用 LLM 提取最近对话的核心主题并生成摘要。
    返回 {theme, summary, full_text} 或 {}。"""
    api, model = _resolve_llm_config(db)
    if not model:
        return {}
    recent = db.execute(
        "SELECT role, content FROM conversations WHERE session_id=? AND expert_id=? ORDER BY id DESC LIMIT 20",
        [session_id, expert_id]
    ).fetchall()
    if not recent:
        return {}
    recent = list(reversed(recent))
    lines = []
    full_lines = []
    for m in recent:
        role_label = "用户" if m["role"] == "user" else "专家"
        lines.append(f"{role_label}: {m['content'][:300]}")
        full_lines.append(f"{role_label}: {m['content']}")
    conversation_text = "\n".join(lines)
    full_text = "\n".join(full_lines)
    prompt = (
        "你是一位知识提炼专家。请从以下对话中提取核心主题和关键信息。\n\n"
        "对话内容：\n" + conversation_text + "\n\n"
        "请用JSON输出，包含两个字段：\n"
        "  - theme: 主题分类，从以下选一个：财富/事业/人性/亲密关系/健康/自我核心/自知/未归类\n"
        "  - summary: 一句话摘要，≤80字，概括这段对话的核心结论或洞察\n\n"
        "仅输出JSON。"
    )
    try:
        result = _call_llm(api, model, [{"role": "user", "content": prompt}])
        content = result.get("content", "").strip()
        if content.startswith("```"):
            content = content.split("\n", 1)[-1]
        if content.endswith("```"):
            content = content.rsplit("```", 1)[0]
        data = json.loads(content)
        return {"theme": data.get("theme", "未归类"), "summary": data.get("summary", ""), "full_text": full_text}
    except Exception as e:
        logger.warning(f"Archive summarize failed: {e}")
        return {}


def _classify_theme(text: str) -> str:
    """本地关键词主题归类"""
    keywords = {
        "财富": ["钱","消费","投资","财务","理财","买","价格","收入"],
        "事业": ["工作","项目","职场","竞争","创业","跳槽", "同事","老板"],
        "人性": ["欲望","嫉妒","人性","动机","渴望","恐惧","羡慕"],
        "亲密关系": ["感情","关系","家庭","伴侣","父母","婚姻","爱"],
        "健康": ["身体","健康","累","疲劳","睡眠","失眠","病","疼"],
        "自我核心": ["我","自己","迷茫","矛盾","意义","价值","改变"],
    }
    for theme, words in keywords.items():
        for w in words:
            if w in text:
                return theme
    return "未归类"


# ═══════════════════════════════════════════════════════════
#  Experts API
# ═══════════════════════════════════════════════════════════

@app.post("/api/experts")
def create_expert(body: dict):
    """创建新专家"""
    db = get_db()
    eid = body.get("id", "").strip().lower().replace(" ", "_")
    if not eid:
        import uuid
        eid = f"custom_{uuid.uuid4().hex[:8]}"
    name = body.get("name", "新专家")
    avatar = body.get("avatar", "🤖")
    domain = body.get("domain", "")
    summon = body.get("summon_phrase", "")
    resp = body.get("response_phrase", "")
    system_prompt = body.get("system_prompt", "")

    # 写入人格文件
    file_path = SKILLS_DIR / "me_experts" / f"{eid}.md"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_text(system_prompt or f"# {name}\n\n待编辑人格设定...", encoding="utf-8")

    # 写入 SQLite
    db.execute(
        """INSERT INTO experts(id,name,avatar,domain,summon_phrase,response_phrase,system_prompt_file)
           VALUES(?,?,?,?,?,?,?)""",
        [eid, name, avatar, domain, summon, resp, str(file_path)]
    )
    db.commit()
    db.close()
    return {"ok": True, "id": eid, "name": name}


@app.get("/api/experts")
def list_experts():
    db = get_db()
    rows = db.execute("SELECT * FROM experts ORDER BY CASE id WHEN 'taishiling' THEN 0 ELSE 1 END, id").fetchall()
    db.close()
    return [ExpertOut(
        id=r["id"], name=r["name"], avatar=r["avatar"], domain=r["domain"],
        summon_phrase=r["summon_phrase"] or "", response_phrase=r["response_phrase"] or "",
        system_prompt="", is_enabled=bool(r["is_enabled"]),
    ).model_dump() for r in rows]

@app.get("/api/experts/{expert_id}")
def get_expert(expert_id: str):
    db = get_db()
    r = db.execute("SELECT * FROM experts WHERE id=?", [expert_id]).fetchone()
    db.close()
    if not r: raise HTTPException(404, "Expert not found")
    prompt = ""
    if r["system_prompt_file"]:
        try: prompt = Path(r["system_prompt_file"]).read_text()
        except: pass
    return ExpertOut(
        id=r["id"], name=r["name"], avatar=r["avatar"], domain=r["domain"],
        summon_phrase=r["summon_phrase"] or "", response_phrase=r["response_phrase"] or "",
        system_prompt=prompt, is_enabled=bool(r["is_enabled"]),
    ).model_dump()

@app.put("/api/experts/{expert_id}")
def update_expert(expert_id: str, body: dict):
    db = get_db()
    existing = db.execute("SELECT id FROM experts WHERE id=?", [expert_id]).fetchone()
    if not existing:
        db.close(); raise HTTPException(404, "Expert not found")
    fields = {k: body.get(k) for k in ["name","avatar","summon_phrase","response_phrase","domain","is_enabled"] if k in body}
    if fields:
        sets = ", ".join(f"{k}=?" for k in fields)
        db.execute(f"UPDATE experts SET {sets} WHERE id=?", list(fields.values())+[expert_id])
    # system_prompt → 写文件
    if body.get("system_prompt"):
        file_path = SKILLS_DIR / "me_experts" / f"{expert_id}.md"
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(body["system_prompt"], encoding="utf-8")
        db.execute("UPDATE experts SET system_prompt_file=? WHERE id=?", [str(file_path), expert_id])
    db.commit(); db.close()
    return {"ok": True}


@app.post("/api/experts/{expert_id}/reset")
def reset_expert(expert_id: str):
    """重置专家到默认值"""
    db = get_db()
    existing = db.execute("SELECT id FROM experts WHERE id=?", [expert_id]).fetchone()
    if not existing:
        db.close(); raise HTTPException(404, "Expert not found")
    defaults = {
        "taishiling": ("太史令","📜","all","史官在侧，秉笔直书。","太史令在此。你的言行，我将如实录于竹帛。",f"{SKILLS_DIR}/me_experts/taishiling.md"),
        "zhongkui":  ("钟馗","⚔️","自我核心","三尺青锋，照我肝胆。","心中有鬼，方须照剑。你是来伏魔的，还是来求饶的？",f"{SKILLS_DIR}/me_experts/zhongkui.md"),
        "chiyou":    ("蚩尤","🐉","事业","兵主旗下，雾散云开。","说敌情。畏刀避剑之人，不配站在我的旗下。",f"{SKILLS_DIR}/me_experts/chiyou.md"),
        "bigan":     ("比干","⚖️","财富","玲珑七窍，公断无私。","我无心，故不偏。把你那笔糊涂账，摊开来。",f"{SKILLS_DIR}/me_experts/bigan.md"),
        "tanlang":   ("贪狼","🐺","人性","贪狼吞月，欲念昭然。","你身上每一寸欲望，都瞒不过我。说吧，这次想喂养哪一个？",f"{SKILLS_DIR}/me_experts/tanlang.md"),
        "zaojun":    ("司命灶君","🔥","亲密关系","灶火明堂，司命在场。","家宅之事，善恶功过，我记下了。从实道来。",f"{SKILLS_DIR}/me_experts/zaojun.md"),
        "qibo":      ("岐伯","🌿","健康","上古天真，问于天师。","身乃心之宅。你哪里失了调和，从实说来。",f"{SKILLS_DIR}/me_experts/qibo.md"),
        "cangjie":     ("仓颉","🏺","自知","鸟兽蹄爪，皆有其迹。","你看到了什么？是河底的石头，还是水面上的波纹？",f"{SKILLS_DIR}/me_experts/cangjie.md"),
    }
    d = defaults.get(expert_id)
    if not d:
        db.close(); raise HTTPException(404, "Not a built-in expert, cannot reset")
    name, avatar, domain, summon, resp, skill_file = d
    db.execute(
        "UPDATE experts SET name=?,avatar=?,domain=?,summon_phrase=?,response_phrase=?,system_prompt_file=?,is_enabled=1 WHERE id=?",
        [name, avatar, domain, summon, resp, skill_file, expert_id]
    )
    db.commit(); db.close()
    return {"ok": True, "message": f"{name} 已恢复默认"}


# ═══════════════════════════════════════════════════════════
#  Knowledge API
# ═══════════════════════════════════════════════════════════

@app.get("/api/knowledge/entries")
def list_knowledge(theme: str = "", search: str = "", limit: int = 50, offset: int = 0):
    db = get_db()
    sql = "SELECT * FROM knowledge_entries WHERE 1=1"
    params = []
    if theme:
        sql += " AND theme_main=?"; params.append(theme)
    if search:
        sql += " AND (summary LIKE ? OR original_text LIKE ? OR theme_main LIKE ?)"
        q = f"%{search}%"; params.extend([q,q,q])
    sql += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    rows = db.execute(sql, params).fetchall()
    db.close()
    return [KnowledgeEntryOut(
        id=r["id"], theme_main=r["theme_main"] or "", theme_sub=r["theme_sub"] or "",
        summary=r["summary"] or "", original_text=r["original_text"] or "",
        expert_id=r["expert_id"] or "", created_at=r["created_at"] or "",
    ).model_dump() for r in rows]

@app.get("/api/knowledge/entries/{entry_id}")
def get_knowledge(entry_id: str):
    db = get_db()
    r = db.execute("SELECT * FROM knowledge_entries WHERE id=?", [entry_id]).fetchone()
    db.close()
    if not r: raise HTTPException(404)
    return KnowledgeEntryOut(
        id=r["id"], theme_main=r["theme_main"] or "", theme_sub=r["theme_sub"] or "",
        summary=r["summary"] or "", original_text=r["original_text"] or "",
        expert_id=r["expert_id"] or "", created_at=r["created_at"] or "",
    ).model_dump()

@app.put("/api/knowledge/entries/{entry_id}")
def update_knowledge(entry_id: str, body: dict):
    db = get_db()
    row = db.execute("SELECT * FROM knowledge_entries WHERE id=?", [entry_id]).fetchone()
    if not row:
        db.close()
        raise HTTPException(404, "entry not found")
    allowed = ["theme_main", "summary", "original_text"]
    sets = []
    params = []
    for k in allowed:
        if k in body:
            sets.append(f"{k}=?")
            params.append(body[k])
    if sets:
        params.append(entry_id)
        db.execute(f"UPDATE knowledge_entries SET {', '.join(sets)} WHERE id=?", params)
        db.commit()
    db.close()
    return {"ok": True}

@app.delete("/api/knowledge/entries/{entry_id}")
def delete_knowledge(entry_id: str):
    db = get_db()
    db.execute("DELETE FROM knowledge_entries WHERE id=?", [entry_id])
    db.commit()
    db.close()
    return {"ok": True}

@app.delete("/api/knowledge/clear")
def clear_knowledge():
    db = get_db()
    db.execute("DELETE FROM knowledge_entries")
    db.commit(); db.close()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════
#  Reports API
# ═══════════════════════════════════════════════════════════

@app.get("/api/reports")
def list_reports():
    """报告列表：优先读 SQLite，回退文件扫描"""
    db = get_db()
    rows = db.execute("SELECT id, filename, period_start, period_end, created_at FROM reports ORDER BY created_at DESC").fetchall()
    db.close()
    if rows:
        return [ReportOut(
            id=r["id"], filename=r["filename"],
            period_start=r["period_start"] or "", period_end=r["period_end"] or "",
            created_at=r["created_at"] or "",
        ).model_dump() for r in rows]
    # fallback: 文件扫描（兼容旧数据）
    reports = []
    if REPORTS_DIR.exists():
        for f in sorted(REPORTS_DIR.glob("*.html"), reverse=True):
            reports.append(ReportOut(
                id=f.stem, filename=f.name,
                created_at=datetime.fromtimestamp(f.stat().st_mtime, tz=timezone.utc).isoformat(),
            ).model_dump())
    return reports

def _markdown_to_html(text: str) -> str:
    """使用 Python markdown 库将 Markdown 转为 HTML"""
    import markdown as md
    return md.markdown(text, extensions=['extra', 'nl2br'])


@app.get("/api/reports/{report_id}")
def get_report(report_id: str):
    # 先从 reports 表查真实文件名
    db = get_db()
    row = db.execute("SELECT filename FROM reports WHERE id=?", [report_id]).fetchone()
    db.close()
    if row:
        path = REPORTS_DIR / row["filename"]
    else:
        # 兼容旧数据
        path = REPORTS_DIR / f"{report_id}.html"
    if not path.exists():
        raise HTTPException(404, "报告文件不存在")
    return FileResponse(path, media_type="text/html; charset=utf-8")


@app.delete("/api/reports/{report_id}")
def delete_report(report_id: str):
    """删除报告：移除 SQLite 记录 + 文件"""
    db = get_db()
    row = db.execute("SELECT filename FROM reports WHERE id=?", [report_id]).fetchone()
    if not row:
        db.close()
        raise HTTPException(404, "报告不存在")
    # 删文件
    path = REPORTS_DIR / row["filename"]
    if path.exists():
        path.unlink()
    # 删数据库记录
    db.execute("DELETE FROM reports WHERE id=?", [report_id])
    db.commit()
    db.close()
    return {"ok": true}

@app.post("/api/reports/generate")
def generate_report():
    """生成综合复盘报告：知识库 + 道痕 + 历史报告 → LLM → HTML"""
    db = get_db()

    # 组装数据
    report_data = _build_report_data(db)
    if report_data["knowledge_total"] == 0 and report_data["daoben_total"] == 0:
        db.close()
        return {"ok": True, "message": "知识库和道痕均为空，无法生成报告"}

    # 调用 LLM
    report_md = _generate_report_analysis(report_data)
    report_html = _markdown_to_html(report_md)

    report_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 生成 HTML
    period_label = f"{report_data['period_start'][:10]} ~ {report_data['period_end'][:10]}"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="utf-8"><title>Me · 复盘报告 {report_id}</title>
<style>
  body {{ font-family: system-ui, -apple-system, sans-serif; max-width: 800px; margin: 0 auto; padding: 40px 20px; color: #1c2e1c; background: #fafcfa; }}
  h1 {{ font-size: 28px; border-bottom: 2px solid #388e3c; padding-bottom: 12px; }}
  h2 {{ font-size: 20px; margin-top: 32px; color: #388e3c; border-left: 4px solid #388e3c; padding-left: 12px; }}
  h3 {{ font-size: 16px; margin-top: 24px; color: #2e7d32; }}
  .report-body {{ line-height: 1.9; }}
  .report-body p {{ margin: 10px 0; }}
  .report-body ul {{ margin: 8px 0; padding-left: 20px; }}
  .report-body ol {{ margin: 8px 0; padding-left: 20px; }}
  .report-body li {{ margin: 6px 0; }}
  .report-body strong {{ color: #1b5e20; }}
  .report-body em {{ color: #558b2f; }}
  .meta {{ font-size: 13px; color: #8a9a8a; margin-bottom: 24px; }}
  .meta span {{ margin-right: 20px; }}
  .footer {{ margin-top: 40px; font-size: 12px; color: #8a9a8a; border-top: 1px solid #d8e4d8; padding-top: 12px; }}
</style></head>
<body>
  <h1>🧬 自知综合评估报告</h1>
  <div class="meta">
    <span>📅 {period_label}</span>
    <span>📚 知识条目 {report_data['knowledge_total']}</span>
    <span>🪨 道痕 {report_data['daoben_total']}</span>
  </div>
  <div class="report-body">
{report_html}
  </div>
  <div class="footer">由 Me · 自知 自动生成 | 基于知识库 + 道痕 + 历史报告综合分析</div>
</body></html>"""

    # 保存
    fname = f"review-{report_id}.html"
    file_path = REPORTS_DIR / fname
    file_path.write_text(html, encoding="utf-8")

    # 写入 reports 表
    db.execute(
        "INSERT INTO reports(id,filename,created_at) VALUES(?,?,?)",
        [report_id, fname, datetime.now(timezone.utc).isoformat()]
    )
    db.commit()
    db.close()

    return {"ok": True, "message": "报告已生成", "report_id": report_id, "filename": fname}


def _build_report_data(db) -> dict:
    """组装报告输入数据：知识库 + 道痕 + 历史报告摘要"""
    now = datetime.now(timezone.utc)
    week_ago = (now - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")

    # 知识库：本周数据
    kb_rows = db.execute(
        "SELECT theme_main, summary, original_text, created_at FROM knowledge_entries ORDER BY created_at DESC LIMIT 200"
    ).fetchall()

    kb_total = len(kb_rows)
    by_theme: dict[str, list] = {}
    for r in kb_rows:
        theme = r["theme_main"] or "未归类"
        if theme not in by_theme:
            by_theme[theme] = []
        by_theme[theme].append({"summary": r["summary"] or "", "text": (r["original_text"] or "")[:300]})

    # 道痕：全部数据
    daoben_rows = db.execute(
        "SELECT main_stone, event_text, greed, fear, excuses, tomorrow_plan, created_at FROM daoben_entries ORDER BY created_at DESC LIMIT 100"
    ).fetchall()

    stone_freq: dict[str, int] = {}
    excuses_freq: dict[str, int] = {}
    greed_count = 0
    fear_count = 0
    for r in daoben_rows:
        s = (r["main_stone"] or "").strip()
        if s:
            stone_freq[s] = stone_freq.get(s, 0) + 1
        e = (r["excuses"] or "").strip()
        if e:
            excuses_freq[e] = excuses_freq.get(e, 0) + 1
        if (r["greed"] or "").strip():
            greed_count += 1
        if (r["fear"] or "").strip():
            fear_count += 1

    top_stones = sorted(stone_freq.items(), key=lambda x: x[1], reverse=True)[:5]
    top_excuses = sorted(excuses_freq.items(), key=lambda x: x[1], reverse=True)[:3]

    # 历史报告摘要
    prev = db.execute(
        "SELECT id, filename, created_at FROM reports ORDER BY created_at DESC LIMIT 2"
    ).fetchall()
    prev_summary = ""
    if len(prev) >= 2:
        prev_path = REPORTS_DIR / prev[1]["filename"]
        if prev_path.exists():
            raw = prev_path.read_text(encoding="utf-8")
            # 提取 body 中的文本摘要（去掉 HTML 标签）
            import re as _re
            text = _re.sub(r'<[^>]+>', ' ', raw)
            text = _re.sub(r'\s+', ' ', text).strip()
            prev_summary = text[:2000]

    return {
        "period_start": week_ago,
        "period_end": now.strftime("%Y-%m-%dT%H:%M:%S"),
        "knowledge_total": kb_total,
        "knowledge_by_theme": {k: {"count": len(v), "samples": [x["summary"] for x in v[:5]]} for k, v in by_theme.items()},
        "daoben_total": len(daoben_rows),
        "top_stones": [{"stone": k, "count": v} for k, v in top_stones],
        "greed_count": greed_count,
        "fear_count": fear_count,
        "top_excuses": [{"excuse": k, "count": v} for k, v in top_excuses],
        "prev_report_summary": prev_summary,
    }


def _generate_report_analysis(report_data: dict) -> str:
    """调用 LLM 生成综合报告。输入结构化数据，输出 Markdown 报告。"""
    api_base, model = _get_llm_config()
    if not model:
        return "（LLM 未配置，无法生成报告）"

    data_json = json.dumps(report_data, ensure_ascii=False, indent=2)

    system_prompt = """你是「自知」平台的深度分析引擎。
自知是一个帮助人观察自己思想河流的系统。用户通过对话、记录道痕、生成知识来留下痕迹，你的任务是把这些痕迹连成镜子——让他看到自己没看到的东西。

## 分析原则
1. 基于证据：每条判断背后要有数据，可以用具体引用
2. 客观评价：不讨好、不安慰、不给鸡汤。直接说观察到的模式
3. 关注模式而非罗列：不要说"讨论了事业3次"，要说"事业讨论里反复出现同一个模式"
4. 有变化才说变化，没变化就说没变化——没有变化本身也是信息
5. 道痕和知识库要交叉验证：如果道痕里的心石和知识库里写的内容有差距，指出来
6. 温度冷静：不煽情、不制造焦虑、也不刻意淡化

## 输出格式

## 画像
一段话，≤200字。基于全部数据评估当前这个人的整体状态。用一段连贯的文字，不单独罗列贪梦和恐惧。结尾给出定性趋势："相较上次，你的总体状态由XX变成了XX"（有上期数据才写这句，没有就跳过）。

## {领域1}
趋势：与上期相比的变化（一两句，没有上期数据则写"无历史对比数据"）
（趋势写完后必须换行）
总结：本期该领域的核心发现
（总结写完后必须换行）
（可自由追加其他有数据支撑的洞察，每个洞察之间也要换行，不要挤成一段）

## {领域2}
...（每个有数据的领域逐一分析）

## 跨域洞察
（如果发现多个领域之间的模式联动，写出来。没有就跳过整段。）

## 风险提示
（如果数据显示某种模式若不干预会走向某处，写出来。没有明确信号就跳过整段。）

注意：
- **领域**是指 `knowledge_by_theme` 里的主题（如"自我核心""事业""自知"），而不是道痕的 stone/excuse/心石。道痕数据（top_stones, top_excuses, greed_count, fear_count）应该融合进画像和跨域洞察，不要单独作为领域章节。
- 每个 knowledge_by_theme 的 key 作为一个领域章节
- 只分析有数据的领域，没数据的不要编造
- 每段结论要有数据支撑
- 画像段是连贯段落，不要用列表或分点
- **格式要求：领域章节内，趋势、总结、洞察每一项之间必须空一行（用双换行分隔），不要把所有内容挤在一个段落里。**"""

    # 明确列出领域，避免 LLM 把道痕数据当领域
    themes = list(report_data.get("knowledge_by_theme", {}).keys())
    theme_list = "\n".join(f"- {t}" for t in themes) if themes else "（无领域数据）"

    user_prompt = f"""以下是本期数据，请按格式生成报告。

**重要：领域章节只分析以下主题（这些是 `knowledge_by_theme` 的 key），不要自行创建其他领域章节。道痕数据（top_stones, top_excuses 等）不是领域，应该融合进画像和跨域洞察里。**

需要分析的领域：
{theme_list}

数据：
```json
{data_json}
```

开始分析。"""

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0.5,
        "max_tokens": 2500,
    }
    try:
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            f"{api_base}/chat/completions", data=body,
            headers={"Content-Type": "application/json"}, method="POST"
        )
        with urllib.request.urlopen(req, timeout=300) as resp:
            data = json.loads(resp.read())
            return data["choices"][0]["message"]["content"]
    except Exception as e:
        logger.warning(f"Report generation failed: {e}")
        return "（LLM 调用失败，请检查配置）"


# ═══════════════════════════════════════════════════════════
#  Settings API
# ═══════════════════════════════════════════════════════════

@app.get("/api/config")
def get_config():
    db = get_db()
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    db.close()
    return {r["key"]: r["value"] for r in rows}

@app.put("/api/config")
def update_config(body: dict):
    db = get_db()
    for k, v in body.items():
        db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", [k, str(v)])
    db.commit()

    # 同步到 me 容器的 config.yaml（embedding 地址）
    if "embedding_api_base" in body:
        _sync_config_yaml("memory.embedding_base_url", body["embedding_api_base"])
    if "embedding_model" in body:
        _sync_config_yaml("memory.embedding_model", body["embedding_model"])

    db.close()
    return {"ok": True}


def _sync_config_yaml(key: str, value: str):
    """将配置同步写入 config.yaml（简约替换）"""
    import yaml
    config_path = ME_HOME / "config.yaml"
    if not config_path.exists():
        return
    try:
        with open(config_path, "r") as f:
            cfg = yaml.safe_load(f) or {}
        # 按点号路径设置值
        keys = key.split(".")
        d = cfg
        for k in keys[:-1]:
            d = d.setdefault(k, {})
        d[keys[-1]] = value
        with open(config_path, "w") as f:
            yaml.dump(cfg, f, default_flow_style=False, allow_unicode=True)
    except Exception:
        pass


@app.get("/api/models")
def list_models(include_embedding: bool = False):
    """从 LLM 接口拉取可用模型列表，兼容 Ollama + OpenAI 两种 API
    
    参数:
        include_embedding: 如果为 True，额外返回 embedding 候选模型（名称含 embed/bge 的）
    """
    db = get_db()
    llm_base = db.execute("SELECT value FROM settings WHERE key='llm_api_base'").fetchone()
    db.close()
    base = (llm_base["value"] if llm_base else "") or LLM_BASE
    base = base.rstrip("/")

    # 1) 先尝试 OpenAI 兼容 API: GET /models
    all_models = []
    try:
        req = urllib.request.Request(f"{base}/models", headers={"Content-Type": "application/json"}, method="GET")
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
            all_models = [m["id"] for m in data.get("data", [])]
    except Exception:
        pass

    # 2) 如果 OpenAI API 没返回，尝试 Ollama API: GET /tags
    if not all_models:
        try:
            req = urllib.request.Request(f"{base}/tags", headers={"Content-Type": "application/json"}, method="GET")
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read())
                all_models = [m["name"] for m in data.get("models", [])]
        except Exception:
            pass

    if not all_models:
        return {
            "ok": True,
            "models": [],
            "embedding_models": ["text-embedding-bge-m3", "bge-large-zh-v1.5", "nomic-embed-text", "text-embedding-3-small"],
            "base": base,
            "error": "无法连接到 LLM 服务，embedding 模型为本地兜底列表",
        }

    # LLM 模型（排除 embedding 专用模型）
    chat_models = [m for m in all_models if "embed" not in m.lower() and "bge" not in m.lower() and "rerank" not in m.lower()]
    # Embedding 候选模型
    embedding_models = [m for m in all_models if "embed" in m.lower() or "bge" in m.lower() or "nomic" in m.lower()]

    return {
        "ok": True,
        "models": chat_models,
        "embedding_models": embedding_models,
        "base": base,
    }


# ═══════════════════════════════════════════════════════════
#  Skills API
# ═══════════════════════════════════════════════════════════

import yaml

def _read_skill_frontmatter(manifest_path: Path) -> dict:
    """从 SKILL.md 的 YAML front matter 提取全部元数据"""
    if not manifest_path.exists():
        return {}
    try:
        content = manifest_path.read_text(encoding="utf-8")
        if content.startswith("---"):
            end = content.find("---", 3)
            if end > 0:
                return yaml.safe_load(content[3:end]) or {}
    except Exception:
        pass
    return {}

def _read_skill_description(manifest_path: Path) -> str:
    """从 SKILL.md 的 YAML front matter 提取 description"""
    return _read_skill_frontmatter(manifest_path).get("description", "")


@app.get("/api/skills")
def list_skills():
    skills = []
    db = get_db()
    db_skills = {r["id"]: r for r in db.execute("SELECT * FROM skills").fetchall()}
    if SKILLS_DIR.exists():
        for d in sorted(SKILLS_DIR.iterdir()):
            if d.is_dir() and not d.name.startswith("."):
                manifest = d / "SKILL.md"
                sid = d.name
                saved = db_skills.get(sid)
                skills.append({
                    "id": sid,
                    "name": sid.replace("_"," ").title(),
                    "has_manifest": manifest.exists(),
                    "description": _read_skill_description(manifest),
                    "enabled": bool(saved["enabled"]) if saved else True,
                })
                # 新发现的 skill 写入 DB
                if not saved:
                    db.execute(
                        "INSERT OR IGNORE INTO skills(id,name,has_manifest,enabled) VALUES(?,?,?,?)",
                        [sid, sid.replace("_"," ").title(), int(manifest.exists()), 1]
                    )
    db.commit(); db.close()
    return skills

@app.put("/api/skills/{skill_id}")
def toggle_skill(skill_id: str, body: dict):
    enabled = body.get("enabled", True)
    db = get_db()
    db.execute(
        "INSERT INTO skills(id,name,has_manifest,enabled) VALUES(?,?,?,?) ON CONFLICT(id) DO UPDATE SET enabled=excluded.enabled",
        [skill_id, skill_id.replace("_"," ").title(), 0, int(enabled)]
    )
    db.commit(); db.close()
    return {"ok": True, "id": skill_id, "enabled": enabled}


@app.delete("/api/skills/{skill_id}")
def delete_skill(skill_id: str):
    """删除技能：移除目录 + SQLite 记录"""
    skill_dir = SKILLS_DIR / skill_id
    if not skill_dir.exists():
        raise HTTPException(404, "技能目录不存在")
    shutil.rmtree(skill_dir)
    db = get_db()
    db.execute("DELETE FROM skills WHERE id=?", [skill_id])
    db.commit()
    db.close()
    return {"ok": True, "id": skill_id}


@app.post("/api/skills/upload")
async def upload_skill_zip(file: UploadFile = File(...)):
    """上传技能压缩包：解压到 SKILLS_DIR，自动注册"""
    if not file.filename or not file.filename.lower().endswith('.zip'):
        raise HTTPException(400, "只支持 .zip 压缩包")

    tmp_path = Path("/tmp") / f"skill_upload_{uuid.uuid4().hex}.zip"
    tmp_path.write_bytes(await file.read())

    db = get_db()
    imported: list[str] = []
    try:
        with zipfile.ZipFile(tmp_path, 'r') as zf:
            for entry in zf.namelist():
                parts = entry.strip('/').split('/')
                if len(parts) < 2:
                    continue  # 跳过根目录文件
                skill_name = parts[0]
                if skill_name.startswith('.') or skill_name.startswith('__'):
                    continue
                inner = '/'.join(parts[1:])
                dest = SKILLS_DIR / skill_name / inner
                dest.parent.mkdir(parents=True, exist_ok=True)
                if entry.endswith('/'):
                    dest.mkdir(parents=True, exist_ok=True)
                else:
                    dest.write_bytes(zf.read(entry))
                    # 发现新技能时写入 DB
                    if skill_name not in imported:
                        manifest = SKILLS_DIR / skill_name / "SKILL.md"
                        db.execute(
                            "INSERT OR IGNORE INTO skills(id,name,has_manifest,enabled) VALUES(?,?,?,?)",
                            [skill_name, skill_name.replace("_"," ").title(), int(manifest.exists()), 1]
                        )
                        imported.append(skill_name)
        db.commit()
    except zipfile.BadZipFile:
        db.close()
        raise HTTPException(400, "无效的 ZIP 文件")
    except Exception as e:
        db.close()
        raise HTTPException(500, f"解压失败: {e}")
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    db.close()
    return {"ok": True, "imported": len(imported), "skills": imported}


# ═══════════════════════════════════════════════════════════
#  File Upload — 账单/文件上传 + 解析
# ═══════════════════════════════════════════════════════════

# 支持的文件类型映射
def _parse_image_with_llm(file_path: Path) -> str:
    """
    使用本地多模态模型识别图片内容。
    将图片转为 base64 → 调用 LLM 的 chat/completions 接口。
    不支持多模态时降级为 OCR / 基本信息描述。
    """
    import base64
    ext = file_path.suffix.lower()
    mime_map = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".webp": "image/webp"}
    mime = mime_map.get(ext, "image/png")

    try:
        image_b64 = base64.b64encode(file_path.read_bytes()).decode()
        data_url = f"data:{mime};base64,{image_b64}"
    except Exception as e:
        return f"[图片: {file_path.name}] 读取失败: {e}"

    prompt = "请识别并描述这张图片中的所有文字和关键信息。如果是账单/收据，请提取：商户名、日期、金额、商品明细。如果是图表，请概括数据趋势。如果只是普通照片，请简要描述内容。"

    # 1) 尝试多模态 LLM
    try:
        api_base, model = _get_llm_config()
        if not model:
            return f"[图片: {file_path.name}] LLM 模型未配置"
        payload = {
            "model": model,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            }],
            "temperature": 0.3,
            "max_tokens": 1024,
        }
        body = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            f"{api_base}/chat/completions", data=body,
            headers={"Content-Type": "application/json"}, method="POST"
        )
        with urllib.request.urlopen(req, timeout=300) as resp:
            data = json.loads(resp.read())
            result = data["choices"][0]["message"]["content"]
            if result and result.strip():
                return result.strip()
    except Exception:
        pass  # 多模态不可用，尝试降级

    # 2) 降级：OCR（tesseract）
    try:
        result = subprocess.run(
            ["tesseract", str(file_path), "stdout", "-l", "chi_sim+eng"],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0 and result.stdout.strip():
            text = result.stdout.strip()[:3000]
            return f"[OCR识别: {file_path.name}]\n{text}"
    except Exception:
        pass

    # 3) 最终降级：返回文件基本信息
    try:
        from PIL import Image
        img = Image.open(file_path)
        size = img.size
        fmt = img.format or "未知"
        return f"[图片: {file_path.name}] 格式={fmt}, 尺寸={size[0]}×{size[1]}, LLM 多模态和 OCR 均不可用。请在系统设置中配置支持多模态的模型或安装 tesseract。"
    except Exception:
        pass

    return f"[图片: {file_path.name}] 无法识别。请确认 LLM 支持多模态或已安装 tesseract。"


FILE_PARSERS = {
    ".pdf":  "pdf",
    ".xlsx": "excel",
    ".xls":  "excel",
    ".csv":  "excel",
    ".png":  "image",
    ".jpg":  "image",
    ".jpeg": "image",
    ".webp": "image",
    ".txt":  "text",
    ".md":   "text",
    ".docx": "word",
}


def _parse_file(file_path: Path) -> str:
    """解析文件内容为文本。按优先级尝试多种方式。"""
    ext = file_path.suffix.lower()
    parser_type = FILE_PARSERS.get(ext, "")

    # 纯文本直接读
    if parser_type == "text" or ext in (".txt", ".md", ".csv"):
        return file_path.read_text(encoding="utf-8", errors="replace")[:MAX_PARSE_CHARS]

    # PDF: 尝试 pdftotext → PyPDF2
    if parser_type == "pdf":
        try:
            result = subprocess.run(
                ["pdftotext", "-layout", "-nopgbrk", str(file_path), "-"],
                capture_output=True, text=True, timeout=15
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout[:MAX_PARSE_CHARS]
        except Exception:
            pass
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(file_path))
            text = "\n".join(p.extract_text() or "" for p in reader.pages)
            return text[:MAX_PARSE_CHARS] if text.strip() else ""
        except Exception:
            pass
        return f"[PDF文件: {file_path.name}] 无法解析文本内容，请确认文件为文本型PDF。"

    # Excel: 尝试 pandas → openpyxl
    if parser_type == "excel":
        try:
            import pandas as pd
            df = pd.read_excel(str(file_path)) if ext in (".xlsx", ".xls") else pd.read_csv(str(file_path))
            return df.to_string(max_rows=9999)[:MAX_PARSE_CHARS]
        except Exception:
            pass
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(file_path), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append(" | ".join(str(c or "") for c in row))
                parts.append(f"=== {sheet_name} ===\n" + "\n".join(rows[:9999]))
            return "\n\n".join(parts)[:MAX_PARSE_CHARS]
        except Exception:
            pass
        return f"[Excel文件: {file_path.name}] 无法解析。"

    # Word
    if parser_type == "word":
        try:
            from docx import Document
            doc = Document(str(file_path))
            return "\n".join(p.text for p in doc.paragraphs)[:MAX_PARSE_CHARS]
        except Exception:
            pass
        return f"[Word文件: {file_path.name}] 无法解析。"

    # 图片: 调用本地多模态模型识别
    if parser_type == "image":
        return _parse_image_with_llm(file_path)

    # 无法识别 → 返回原始文本尝试
    try:
        return file_path.read_text(encoding="utf-8", errors="replace")[:3000]
    except Exception:
        return f"[文件: {file_path.name}] 无法识别格式。"


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    上传文件 → 保存 + 解析文本内容 → 立即返回。
    前端收到解析内容后作为 chat 消息发送给专家。
    """
    saved_name = file.filename or "upload"
    saved_path = UPLOAD_DIR / saved_name
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    # 重名时自动加序号
    if saved_path.exists():
        stem = saved_path.stem
        ext = saved_path.suffix
        i = 1
        while (UPLOAD_DIR / f"{stem} ({i}){ext}").exists():
            i += 1
        saved_name = f"{stem} ({i}){ext}"
        saved_path = UPLOAD_DIR / saved_name
    with saved_path.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    parsed_text = _parse_file(saved_path)
    return {
        "ok": True,
        "filename": file.filename,
        "saved_as": saved_name,
        "file_type": FILE_PARSERS.get(ext, "unknown"),
        "parsed_length": len(parsed_text),
        "parsed_text": parsed_text[:MAX_PARSE_CHARS],
        "truncated": len(parsed_text) > MAX_PARSE_CHARS,
    }


@app.get("/api/uploads/{filename}")
def get_uploaded_file(filename: str):
    """获取已上传的文件"""
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise HTTPException(404)
    return FileResponse(path)


@app.get("/api/uploads")
def list_uploads():
    """列出所有已上传的文件"""
    if not UPLOAD_DIR.exists():
        return {"files": []}
    files = []
    for f in sorted(UPLOAD_DIR.iterdir(), key=lambda x: x.stat().st_mtime, reverse=True):
        if f.is_file():
            st = f.stat()
            files.append({
                "name": f.name,
                "size": st.st_size,
                "mtime": datetime.fromtimestamp(st.st_mtime, tz=timezone.utc).isoformat(),
                "ext": f.suffix.lower(),
            })
    return {"files": files}


@app.delete("/api/uploads/{filename}")
def delete_uploaded_file(filename: str):
    """删除已上传的文件"""
    path = UPLOAD_DIR / filename
    if not path.exists():
        raise HTTPException(404)
    path.unlink()
    return {"ok": True}


# ═══════════════════════════════════════════════════════════
#  道痕日记 (Daoben Diary) API
# ═══════════════════════════════════════════════════════════

@app.get("/api/daoben/entries")
def list_daoben(search: str = "", limit: int = 50, offset: int = 0):
    """列出道痕条目，支持搜索"""
    db = get_db()
    sql = "SELECT * FROM daoben_entries WHERE 1=1"
    params = []
    if search:
        sql += " AND (event_text LIKE ? OR main_stone LIKE ? OR greed LIKE ? OR fear LIKE ?)"
        q = f"%{search}%"
        params.extend([q, q, q, q])
    sql += " ORDER BY created_at DESC LIMIT ? OFFSET ?"
    params.extend([limit, offset])
    rows = db.execute(sql, params).fetchall()
    db.close()
    return [DaobenEntryOut(
        id=r["id"], event_text=r["event_text"] or "", first_reaction=r["first_reaction"] or "",
        greed=r["greed"] or "", fear=r["fear"] or "", excuses=r["excuses"] or "",
        main_stone=r["main_stone"] or "", tomorrow_plan=r["tomorrow_plan"] or "",
        expert_id=r["expert_id"] or "", source=r["source"] or "manual",
        created_at=r["created_at"] or "",
    ).model_dump() for r in rows]


@app.get("/api/daoben/entries/{entry_id}")
def get_daoben_entry(entry_id: str):
    """获取单条道痕"""
    db = get_db()
    r = db.execute("SELECT * FROM daoben_entries WHERE id=?", [entry_id]).fetchone()
    db.close()
    if not r:
        raise HTTPException(404, "道痕条目不存在")
    return DaobenEntryOut(
        id=r["id"], event_text=r["event_text"] or "", first_reaction=r["first_reaction"] or "",
        greed=r["greed"] or "", fear=r["fear"] or "", excuses=r["excuses"] or "",
        main_stone=r["main_stone"] or "", tomorrow_plan=r["tomorrow_plan"] or "",
        expert_id=r["expert_id"] or "", source=r["source"] or "manual",
        created_at=r["created_at"] or "",
    ).model_dump()


@app.post("/api/daoben/entries")
def create_daoben_entry(body: dict):
    """创建道痕条目（手动或 cangjie 对话后自动入库）"""
    import uuid as _uuid
    entry_id = body.get("id") or f"daoben_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S%f')}"
    db = get_db()
    db.execute(
        """INSERT INTO daoben_entries(id,event_text,first_reaction,greed,fear,excuses,main_stone,tomorrow_plan,expert_id,source)
           VALUES(?,?,?,?,?,?,?,?,?,?)""",
        [
            entry_id,
            body.get("event_text", ""),
            body.get("first_reaction", ""),
            body.get("greed", ""),
            body.get("fear", ""),
            body.get("excuses", ""),
            body.get("main_stone", ""),
            body.get("tomorrow_plan", ""),
            body.get("expert_id", ""),
            body.get("source", "manual"),
        ]
    )
    db.commit()
    db.close()
    return {"ok": True, "id": entry_id}


@app.delete("/api/daoben/entries/{entry_id}")
def delete_daoben_entry(entry_id: str):
    """删除单条道痕"""
    db = get_db()
    db.execute("DELETE FROM daoben_entries WHERE id=?", [entry_id])
    db.commit()
    db.close()
    return {"ok": True}


@app.get("/api/daoben/dashboard")
def daoben_dashboard(period: str = "week"):
    """道痕回看仪表盘：周期回顾（week/month/all）
    返回：累计统计、重复石头、贪惧天平、借口排行、趋势数据"""
    db = get_db()
    now = datetime.now(timezone.utc)
    if period == "week":
        since = (now - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%S")
    elif period == "month":
        since = (now - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%S")
    else:
        since = None

    base_sql = "SELECT * FROM daoben_entries"
    params = []
    if since:
        base_sql += " WHERE created_at >= ?"
        params.append(since)
    base_sql += " ORDER BY created_at DESC"

    rows = db.execute(base_sql, params).fetchall()
    db.close()

    total = len(rows)
    if total == 0:
        return {"total": 0, "stones": [], "greed_fear_ratio": None, "excuses": [], "timeline": [], "message": "该时段暂无道痕记录"}

    # 重复石头统计
    stone_count = {}
    for r in rows:
        s = (r["main_stone"] or "").strip()
        if s:
            stone_count[s] = stone_count.get(s, 0) + 1
    stones = sorted([{"stone": k, "count": v} for k, v in stone_count.items()], key=lambda x: x["count"], reverse=True)[:10]

    # 贪/惧天平（有内容的比例）
    has_greed = sum(1 for r in rows if (r["greed"] or "").strip())
    has_fear = sum(1 for r in rows if (r["fear"] or "").strip())
    greed_fear_ratio = round(has_greed / max(has_fear, 1), 2)

    # 借口模式 Top 5
    excuse_count = {}
    for r in rows:
        e = (r["excuses"] or "").strip()
        if e:
            excuse_count[e] = excuse_count.get(e, 0) + 1
    excuses = sorted([{"excuse": k, "count": v} for k, v in excuse_count.items()], key=lambda x: x["count"], reverse=True)[:5]

    # 时间线（按天聚合）
    daily = {}
    for r in rows:
        day = (r["created_at"] or "")[:10]
        if day:
            daily[day] = daily.get(day, 0) + 1
    timeline = [{"date": k, "count": v} for k, v in sorted(daily.items())]

    # 趋势判断
    trend = "stable"
    if len(stones) >= 2:
        top_stone_pct = stones[0]["count"] / total
        if top_stone_pct > 0.5:
            trend = "dominated"
        elif top_stone_pct > 0.3:
            trend = "skewed"

    return {
        "total": total,
        "period": period,
        "stones": stones,
        "greed_fear_ratio": greed_fear_ratio,
        "greed_fear_label": "贪多惧少" if greed_fear_ratio > 1.5 else ("惧多贪少" if greed_fear_ratio < 0.67 else "贪惧均衡"),
        "excuses": excuses,
        "timeline": timeline,
        "trend": trend,
        "trend_label": {"stable": "心石分布均匀", "skewed": "某类心石偏高", "dominated": "单一心石主导"}.get(trend, ""),
    }


@app.get("/api/daoben/stats")
def daoben_stats():
    """道痕统计：重复石头 top5"""
    db = get_db()
    rows = db.execute(
        "SELECT main_stone, COUNT(*) as cnt FROM daoben_entries WHERE main_stone != '' GROUP BY main_stone ORDER BY cnt DESC LIMIT 5"
    ).fetchall()
    db.close()
    return [{"stone": r["main_stone"], "count": r["cnt"]} for r in rows]


# ═══════════════════════════════════════════════════════════
#  Export API
# ═══════════════════════════════════════════════════════════

@app.get("/api/export")
def export_data():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # SQLite dump
        zf.writestr("me.db", DB_PATH.read_bytes() if DB_PATH.exists() else b"")
        # Reports
        if REPORTS_DIR.exists():
            for f in REPORTS_DIR.rglob("*.html"):
                zf.write(f, f"reports/{f.name}")
        # Knowledge entries JSON
        db = get_db()
        rows = db.execute("SELECT * FROM knowledge_entries").fetchall()
        zf.writestr("knowledge_export.json", json.dumps([dict(r) for r in rows], ensure_ascii=False, indent=2))
        # Daoben entries JSON
        daoben_rows = db.execute("SELECT * FROM daoben_entries").fetchall()
        zf.writestr("daoben_export.json", json.dumps([dict(r) for r in daoben_rows], ensure_ascii=False, indent=2))
        db.close()
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=me-export-{datetime.now().strftime('%Y%m%d')}.zip"}
    )


# ─── Entry ─────────────────────────────────────────────────

# 定时复盘后台线程
_scheduler_stop = threading.Event()

def _scheduler_loop():
    """每 10 分钟检查一次：周日 20:00 生成周报，每月 1 日 20:00 生成月报"""
    while not _scheduler_stop.wait(600):  # 10 分钟
        try:
            now = datetime.now()
            # 检查最近一次生成的报告类型和时间
            db = get_db()
            last = db.execute(
                "SELECT created_at, filename FROM reports ORDER BY created_at DESC LIMIT 1"
            ).fetchone()
            db.close()

            # 周日 20:00-20:10 间检查周报
            if now.weekday() == 6 and now.hour == 20 and now.minute < 10:
                needed = True
                if last:
                    last_dt = datetime.fromisoformat(last["created_at"].replace("Z", "+00:00"))
                    if last_dt.date() == now.date():
                        needed = False  # 今天已生成过
                if needed:
                    _do_scheduled_report("weekly")

            # 每月 1 日 20:00-20:10 间检查月报
            if now.day == 1 and now.hour == 20 and now.minute < 10:
                needed = True
                if last:
                    last_dt = datetime.fromisoformat(last["created_at"].replace("Z", "+00:00"))
                    if last_dt.month == now.month and last_dt.year == now.year:
                        needed = False
                if needed:
                    _do_scheduled_report("monthly")
        except Exception:
            pass


def _do_scheduled_report(kind: str):
    """触发定时报告生成（复用统一报告管道）"""
    db = get_db()
    report_data = _build_report_data(db)
    if report_data["knowledge_total"] == 0 and report_data["daoben_total"] == 0:
        db.close()
        return

    report_md = _generate_report_analysis(report_data)
    report_html = _markdown_to_html(report_md)

    label = "周报" if kind == "weekly" else "月报"
    report_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    period_label = f"{report_data['period_start'][:10]} ~ {report_data['period_end'][:10]}"

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="utf-8"><title>Me · {label} {report_id}</title>
<style>
  body {{ font-family: system-ui, -apple-system, sans-serif; max-width: 800px; margin: 0 auto; padding: 40px 20px; color: #1c2e1c; background: #fafcfa; }}
  h1 {{ font-size: 28px; border-bottom: 2px solid #388e3c; padding-bottom: 12px; }}
  h2 {{ font-size: 20px; margin-top: 32px; color: #388e3c; border-left: 4px solid #388e3c; padding-left: 12px; }}
  h3 {{ font-size: 16px; margin-top: 24px; color: #2e7d32; }}
  .report-body {{ line-height: 1.9; }}
  .report-body p {{ margin: 10px 0; }}
  .report-body ul {{ margin: 8px 0; padding-left: 20px; }}
  .report-body ol {{ margin: 8px 0; padding-left: 20px; }}
  .report-body li {{ margin: 6px 0; }}
  .report-body strong {{ color: #1b5e20; }}
  .meta {{ font-size: 13px; color: #8a9a8a; margin-bottom: 24px; }}
  .meta span {{ margin-right: 20px; }}
  .footer {{ margin-top: 40px; font-size: 12px; color: #8a9a8a; border-top: 1px solid #d8e4d8; padding-top: 12px; }}
</style></head>
<body>
  <h1>🧬 自知{label}</h1>
  <div class="meta">
    <span>📅 {period_label}</span>
    <span>📚 知识条目 {report_data['knowledge_total']}</span>
    <span>🪨 道痕 {report_data['daoben_total']}</span>
  </div>
  <div class="report-body">
{report_html}
  </div>
  <div class="footer">由 Me · 自知 自动生成</div>
</body></html>"""

    fname = f"{kind}-{report_id}.html"
    file_path = REPORTS_DIR / fname
    file_path.write_text(html, encoding="utf-8")

    db.execute(
        "INSERT INTO reports(id,filename,created_at) VALUES(?,?,?)",
        [report_id, fname, datetime.now(timezone.utc).isoformat()]
    )
    db.commit()
    db.close()


# 启动后台定时器
_scheduler_thread = threading.Thread(target=_scheduler_loop, daemon=True)
_scheduler_thread.start()

# ═══════════════════════════════════════════════════════════
#  Static Files（Vue PWA 前端）
# ═══════════════════════════════════════════════════════════
STATIC_DIR = Path(os.environ.get("ME_STATIC_DIR", "/app/static"))
if STATIC_DIR.exists():
    # SPA fallback：非 /api 路径 → index.html
    @app.get("/{full_path:path}")
    async def spa_fallback(full_path: str):
        file_path = STATIC_DIR / full_path
        if file_path.is_file():
            return FileResponse(file_path)
        return FileResponse(STATIC_DIR / "index.html")

    # 根路径
    @app.get("/")
    async def root():
        return FileResponse(STATIC_DIR / "index.html")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=6868)
